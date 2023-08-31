from selenium import webdriver
import openpyxl
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,rownum,columnno,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)
class Login:
    # text_field_username_xpath='//input[@formcontrolname="email"]'
    # text_field_password_name="password"
    # button_login_xpath='//button[@type="submit"]'

    def __init__(self,driver):
        self.driver=driver

    def setUserName(self,username):
        sh=self.driver.find_element('xpath',"//input[@formcontrolname='email']")
        sh.clear()
        # self.driver.find_element(self.text_field_username_xpath).send_keys("haritha.k@testyantra.in")
        sh.send_keys("haritha.k@testyantra.in")


    def setPassword(self,password):
        self.driver.find_element('name','password').clear()
        self.driver.find_element('name','password').send_keys("9346144936")

    def clickLogin(self):
        self.driver.find_element('xpath','//button[@type="submit"]').click()
    #
    def clickLogout(self):
        self.driver.find_element('xpath','//button[@id="dropdownMenuButton1"]').click()
        self.driver.find_element('xpath','//a[text()="Logout"]').click()