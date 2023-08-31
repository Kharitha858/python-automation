# # Import the necessary libraries
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service as ChromeService
# from webdriver_manager.chrome import ChromeDriverManager
# from selenium.webdriver.common.by import By
# import time
#
# # Install the ChromeDriver executable and start a Chrome browser using Selenium
# driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
#
# # Navigate to the website and maximize the window
# driver.get('http://10.10.20.23:3001/')
# driver.maximize_window()
# driver.find_element(By.XPATH,'//p[text()="Login/SignUp"]').click()
# time.sleep(2)
# driver.find_element(By.XPATH,'//span[text()=" Login as Admin"]').click()
# driver.find_element(By.XPATH,'//input[@placeholder="Your Username*"]').send_keys("Raunak")
# driver.find_element(By.XPATH,'//input[@placeholder="Your Password"]').send_keys("1234")
# # time.sleep(3)
# driver.find_element(By.XPATH,'//button[text()=" Login"]').click()
# time.sleep(2)
# driver.find_element(By.XPATH,'//div[@class="ant-layout-sider-children"]//descendant::p[text()="Manage Master"]//ancestor::li').click()
# time.sleep(2)
# driver.find_element(By.XPATH,'//ul[@class="ant-menu ant-menu-sub ant-menu-vertical"]//p[text()="  Basic Settings"]').click()
# time.sleep(2)
# driver.find_element(By.XPATH,'//ul[@class="ant-menu ant-menu-sub ant-menu-vertical"]//p[text()="Designation"]').click()
# time.sleep(2)
#
# # @@@@@@@@@@@@ using display dropdown:
#
# # d=driver.find_element(By.XPATH,'//div[@class="ant-select-selector"]').click()
# # display=driver.find_elements(By.XPATH,'//div[@class="ant-select-item ant-select-item-option"]')
# # # display.click()
# # time.sleep(3)
# # for r in display:
# #     time.sleep(3)
# #     print(r.text)
# #     time.sleep(2)
# #     if r.text=='100':
# #         r.click()
# #         time.sleep(3)
# #
# # # scroll=driver.execute_script("window.scrollBy(0,2000)","")
# # # time.sleep(3)
# # data=driver.find_elements(By.XPATH,'//div[@class="ant-table-body"]//tr//td[2]')
# # for desg in data:
# #     time.sleep(2)
# #     print(desg.text)
#
#
# # @@@@@@@@@@@@@@ using pagination:
# pagenation=driver.find_elements(By.XPATH,'//li[contains(@class,"ant-pagination-item ant-pagination-item")]')
# for page in pagenation:
#     page.click()
#     time.sleep(2)
#     data = driver.find_elements(By.XPATH, '//div[@class="ant-table-body"]//tr//td[2]')
#     for desg in data:
#         time.sleep(2)
#         data=desg.text
#         print(data)


#working with excel sheet:

# import openpyxl
# path=r"C:\Users\CBT\PycharmProjects\GOS\python.xlsx"
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(row=r,column=c).value="welcome"
# workbook.save(path)


import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import time

# Install the ChromeDriver executable and start a Chrome browser using Selenium
driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

# Navigate to the website and maximize the window
driver.get('https://godsownstay.com/bookings')
driver.maximize_window()
driver.find_element(By.XPATH,' //a[text()=" Login / "]').click()
time.sleep(2)
driver.find_element(By.XPATH,' //a[text()="Login as Admin"]').click()
driver.find_element(By.XPATH,'//input[@id="username"]').send_keys("Harshith007")
driver.find_element(By.XPATH,'//input[@id="password"]').send_keys("Harshith@123")
# time.sleep(3)
driver.find_element(By.XPATH,'//button[@class="btnSubmit    btn-block"]').click()
time.sleep(3)
driver.find_element(By.XPATH,'//i[@class="menu-icon fa fa-home"]').click()
time.sleep(2)
driver.find_element(By.XPATH,'//a[text()=" Basic Settings"]').click()
# time.sleep(2)
driver.find_element(By.XPATH,'//a[text()=" Designation"]').click()
time.sleep(2)

pagenation=driver.find_elements(By.XPATH,'//div[@id="dynamic-table_paginate"]//li[contains(@class,"paginate_button active")]')
for page in pagenation:
    # page.click()
    time.sleep(2)
    data = driver.find_elements(By.XPATH, '//table[@id="dynamic-table"]//tbody//tr')
    for desg in data:
        time.sleep(2)
        data=desg.text
        path = r"C:\Users\CBT\PycharmProjects\GOS\python.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        for r in range(1, 11):
            # print(r)
            for c in range(1, 2):
                # time.sleep(2)
                # print(c)
                sheet.cell(row=r, column=c).value = data
        workbook.save(path)
        print(data)























